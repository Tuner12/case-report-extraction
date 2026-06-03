#!/usr/bin/env python3
"""Heuristically audit whether a case workbook or JSON is supported by source PDF text."""

from __future__ import annotations

import argparse
import json
import re
from pathlib import Path

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover - optional when auditing JSON only.
    load_workbook = None


STOPWORDS = {
    "about",
    "after",
    "also",
    "and",
    "because",
    "been",
    "being",
    "case",
    "could",
    "from",
    "have",
    "into",
    "malignant",
    "melanoma",
    "patient",
    "patients",
    "report",
    "should",
    "that",
    "the",
    "then",
    "there",
    "this",
    "through",
    "with",
    "were",
    "what",
    "when",
    "will",
}


def normalize(text: str) -> str:
    return re.sub(r"[^a-z0-9]+", " ", text.lower()).strip()


def content_tokens(text: str) -> list[str]:
    return [
        token
        for token in normalize(text).split()
        if len(token) >= 4 and token not in STOPWORDS and not token.isdigit()
    ]


def consultation_blocks(data: dict) -> list[tuple[str, dict]]:
    blocks = []
    for key, value in data.items():
        if re.fullmatch(r"Cons(?:u)?lation\d+", key) and isinstance(value, dict):
            blocks.append((key, value))
    return sorted(blocks, key=lambda item: int(re.search(r"\d+", item[0]).group(0)))


def fields_from_json(path: Path) -> tuple[str | None, list[tuple[str, str]]]:
    data = json.loads(path.read_text(encoding="utf-8"))
    fields = []
    for key, block in consultation_blocks(data):
        for field in ("Record1", "Answer1"):
            fields.append((f"{key}.{field}", str(block.get(field) or "")))
    return data.get("CRID"), fields


def fields_from_workbook(path: Path) -> tuple[str | None, list[tuple[str, str]]]:
    if load_workbook is None:
        raise RuntimeError("openpyxl is required to audit .xlsx workbooks")
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
    values = [ws.cell(2, col).value for col in range(1, ws.max_column + 1)]
    case_id = values[0] if values else None
    fields = []
    for index, header in enumerate(headers):
        if not header:
            continue
        header_text = str(header).strip()
        value = str(values[index] or "")
        if re.fullmatch(r"Record\s+\d+", header_text, flags=re.IGNORECASE):
            fields.append((header_text, value))
        elif re.fullmatch(r"Answer\s+\d+", header_text, flags=re.IGNORECASE):
            fields.append((header_text, value))
        elif re.fullmatch(r"Anwser\s+\d+", header_text, flags=re.IGNORECASE):
            fields.append((header_text, value))
        elif re.fullmatch(r"Patient record\s+\d+", header_text, flags=re.IGNORECASE):
            fields.append((header_text, value))
        elif re.fullmatch(r"Clinician recommendation\s+\d+", header_text, flags=re.IGNORECASE):
            fields.append((header_text, value))
        elif header_text.lower() in {"final follow up", "final output"}:
            fields.append((header_text, value))
    return str(case_id) if case_id else None, fields


def fields_from_input(path: Path) -> tuple[str | None, list[tuple[str, str]]]:
    suffix = path.suffix.lower()
    if suffix == ".json":
        return fields_from_json(path)
    if suffix in {".xlsx", ".xlsm"}:
        return fields_from_workbook(path)
    raise RuntimeError(f"Unsupported input type for source audit: {path.suffix}")


def phrase_hits(field_text: str, source_norm: str, size: int = 6) -> int:
    words = normalize(field_text).split()
    if len(words) < size:
        return 0
    hits = 0
    for index in range(0, len(words) - size + 1):
        phrase = " ".join(words[index : index + size])
        if phrase in source_norm:
            hits += 1
    return hits


def field_report(name: str, text: str, source_tokens: set[str], source_norm: str) -> dict:
    tokens = content_tokens(text)
    if not tokens:
        return {"field": name, "status": "empty", "coverage": 1.0, "phrase_hits": 0, "missing": []}
    unique = []
    for token in tokens:
        if token not in unique:
            unique.append(token)
    covered = [token for token in unique if token in source_tokens]
    missing = [token for token in unique if token not in source_tokens]
    coverage = len(covered) / len(unique)
    hits = phrase_hits(text, source_norm)
    status = "ok"
    if coverage < 0.45 and hits == 0:
        status = "fail"
    elif coverage < 0.6 and hits == 0:
        status = "warn"
    return {
        "field": name,
        "status": status,
        "coverage": round(coverage, 3),
        "phrase_hits": hits,
        "missing": missing[:20],
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("input_file", type=Path, help="Case .xlsx workbook or legacy .json file")
    parser.add_argument("--source-text", type=Path, required=True)
    parser.add_argument("--write-report", type=Path)
    parser.add_argument("--fail-on-warn", action="store_true")
    parser.add_argument(
        "--include-questions",
        action="store_true",
        help="Also audit generated clinical questions in legacy JSON inputs. Workbook questions are skipped by default.",
    )
    args = parser.parse_args()

    case_id, extracted_fields = fields_from_input(args.input_file)
    source = args.source_text.read_text(encoding="utf-8", errors="ignore")
    source_norm = normalize(source)
    source_tokens = set(source_norm.split())

    report = {
        "input_file": str(args.input_file),
        "source_text": str(args.source_text),
        "CRID": case_id,
        "fields": [],
        "warnings": [],
        "errors": [],
    }

    if args.input_file.suffix.lower() == ".json" and args.include_questions:
        data = json.loads(args.input_file.read_text(encoding="utf-8"))
        extracted_fields = []
        for key, block in consultation_blocks(data):
            for field in ("Record1", "Question1", "Answer1"):
                extracted_fields.append((f"{key}.{field}", str(block.get(field) or "")))

    for field_name, text in extracted_fields:
        item = field_report(field_name, text, source_tokens, source_norm)
        report["fields"].append(item)
        if item["status"] == "warn":
            report["warnings"].append(f"{item['field']} has weak source support")
        elif item["status"] == "fail":
            report["errors"].append(f"{item['field']} appears unsupported by source text")

    if args.write_report:
        args.write_report.parent.mkdir(parents=True, exist_ok=True)
        args.write_report.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")

    print(json.dumps(report, ensure_ascii=False, indent=2))
    if report["errors"] or (args.fail_on_warn and report["warnings"]):
        raise SystemExit(1)


if __name__ == "__main__":
    main()
