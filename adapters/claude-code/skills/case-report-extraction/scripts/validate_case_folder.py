#!/usr/bin/env python3
"""Validate a standard case-report extraction folder."""

from __future__ import annotations

import argparse
import json
import re
from pathlib import Path

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover - optional when validating folders only.
    load_workbook = None

try:
    import fitz
except Exception:  # pragma: no cover - optional PDF annotation validation.
    fitz = None


def consultation_blocks(data: dict) -> list[tuple[str, dict]]:
    blocks = []
    for key, value in data.items():
        if re.fullmatch(r"Cons(?:u)?lation\d+", key) and isinstance(value, dict):
            blocks.append((key, value))
    return sorted(blocks, key=lambda item: int(re.search(r"\d+", item[0]).group(0)))


def is_table_workbook(path: Path) -> bool:
    return re.search(r"_table\d+\.xlsx$", path.name, flags=re.IGNORECASE) is not None


def split_resource_names(value) -> list[str]:
    if value is None or value == "":
        return []
    names = []
    for item in str(value).split(","):
        item = item.strip()
        for name in re.findall(r"[\w.-]+\.(?:png|jpg|jpeg|txt|xlsx|xlsm|pdf)", item, flags=re.IGNORECASE):
            names.append(Path(name).name)
    return names


def count_highlight_annotations(pdf: Path) -> int | None:
    if fitz is None:
        return None
    doc = fitz.open(pdf)
    count = 0
    for page in doc:
        annot = page.first_annot
        while annot:
            if annot.type and annot.type[1] == "Highlight":
                count += 1
            annot = annot.next
    doc.close()
    return count


def validate_evidence_pdf(folder: Path, case_id: str, errors: list[str], warnings: list[str]) -> dict:
    expected = folder / f"{case_id}.pdf"
    report = {"evidence_pdf": str(expected)}
    if not expected.exists():
        errors.append(f"Missing final evidence-highlighted PDF: {case_id}.pdf")
        return report
    highlight_count = count_highlight_annotations(expected)
    report["highlight_annotations"] = highlight_count
    if highlight_count is None:
        warnings.append("PyMuPDF unavailable; PDF highlight annotations not inspected")
    elif highlight_count == 0:
        errors.append(f"Final PDF has no highlight annotations: {case_id}.pdf")
    if not (folder / "evidence_highlight_report.json").exists():
        warnings.append("Missing working evidence_highlight_report.json")
    return report


def validate_workbook(folder: Path, workbook: Path | None, errors: list[str], warnings: list[str]) -> dict:
    report = {}
    if workbook is None:
        workbooks = [p for p in sorted(folder.glob("*.xlsx")) if not is_table_workbook(p)]
        workbook = workbooks[0] if workbooks else None
    if workbook is None:
        warnings.append("No main case workbook found")
        return report
    report["workbook"] = str(workbook)
    if load_workbook is None:
        warnings.append("openpyxl unavailable; workbook contents not inspected")
        return report
    wb = load_workbook(workbook, data_only=True)
    ws = wb.active
    headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
    values = [ws.cell(2, col).value for col in range(1, ws.max_column + 1)]
    report["CRID_from_workbook"] = values[0] if values else None
    if not values or not values[0]:
        errors.append("Workbook missing CRID in row 2 column A")
    record_count = sum(
        1 for header in headers if header and re.fullmatch(r"(Record|Patient record)\s+\d+", str(header), flags=re.IGNORECASE)
    )
    report["record_count"] = record_count
    if record_count == 0:
        errors.append("Workbook has no Record N or Patient record N columns")
    for row_number, kind in ((3, "figure"), (4, "table")):
        for col in range(2, ws.max_column + 1):
            for name in split_resource_names(ws.cell(row_number, col).value):
                if not (folder / name).exists():
                    errors.append(f"Workbook references missing {kind}: {name}")
    return report


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("case_folder", type=Path)
    parser.add_argument("--json", dest="json_file", type=Path, default=None)
    parser.add_argument("--workbook", type=Path, default=None)
    parser.add_argument("--write-report", action="store_true")
    args = parser.parse_args()

    folder = args.case_folder.expanduser().resolve()
    json_file = args.json_file

    errors = []
    warnings = []
    files = [p.name for p in folder.iterdir()] if folder.exists() else []
    report = {"case_folder": str(folder), "files": files, "errors": errors, "warnings": warnings}

    if not folder.exists():
        errors.append("Case folder does not exist")
    if folder.exists():
        report.update(validate_workbook(folder, args.workbook, errors, warnings))
        case_id = str(report.get("CRID_from_workbook") or folder.name)
        report.update(validate_evidence_pdf(folder, case_id, errors, warnings))

    if json_file and json_file.exists():
        data = json.loads(json_file.read_text(encoding="utf-8"))
        report["CRID"] = data.get("CRID")
        blocks = consultation_blocks(data)
        report["consultation_count"] = len(blocks)
        if not data.get("CRID"):
            errors.append("JSON missing CRID")
        if not blocks:
            errors.append("JSON has no ConsulationN blocks")
        last_key = blocks[-1][0] if blocks else None
        has_final_follow_up = bool(
            data.get("FinalFollowUp") or data.get("Final follow up") or data.get("Final output")
        )
        for key, block in blocks:
            if not block.get("Record1"):
                errors.append(f"{key} missing Record1")
            if not block.get("Answer1") and not block.get("FinalFollowUp") and not (
                key == last_key and has_final_follow_up
            ):
                warnings.append(f"{key} missing Answer1")
            for kind in ("Figures", "Tables"):
                for item in block.get(kind, []) or []:
                    if isinstance(item, dict):
                        name = item.get("Name") or Path(str(item.get("Path", ""))).name
                    else:
                        name = str(item)
                    if name and not (folder / name).exists():
                        errors.append(f"{key} references missing {kind[:-1].lower()}: {name}")
    elif args.json_file:
        errors.append("Requested JSON file does not exist")

    if args.write_report and folder.exists():
        (folder / "validation_report.json").write_text(
            json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8"
        )
    print(json.dumps(report, ensure_ascii=False, indent=2))
    raise SystemExit(1 if errors else 0)


if __name__ == "__main__":
    main()
