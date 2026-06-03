#!/usr/bin/env python3
"""Create a sentence/keyword-level evidence-highlighted PDF."""

from __future__ import annotations

import argparse
import json
import math
import re
from dataclasses import dataclass
from pathlib import Path

import fitz
from openpyxl import load_workbook


STOPWORDS = {
    "about",
    "after",
    "also",
    "because",
    "been",
    "being",
    "case",
    "could",
    "from",
    "have",
    "into",
    "management",
    "malignant",
    "melanoma",
    "patient",
    "patients",
    "report",
    "should",
    "that",
    "there",
    "this",
    "through",
    "treatment",
    "with",
    "were",
    "what",
    "when",
    "will",
}

FIELD_RE = re.compile(
    r"^(Record|Answer|Anwser|Patient record|Clinician recommendation)\s+\d+$",
    flags=re.IGNORECASE,
)
FINAL_FIELDS = {"final follow up", "final output"}
COLORS = {
    "record": (1.0, 0.86, 0.15),
    "answer": (0.45, 0.90, 0.45),
    "final": (1.0, 0.62, 0.25),
}


@dataclass
class Field:
    name: str
    text: str
    kind: str


@dataclass
class Block:
    page_index: int
    rect: fitz.Rect
    text: str
    tokens: set[str]
    norm: str


@dataclass
class Word:
    page_index: int
    rect: fitz.Rect
    text: str
    tokens: set[str]
    block_no: int
    line_no: int
    word_no: int


@dataclass
class Sentence:
    page_index: int
    words: list[Word]
    text: str
    tokens: set[str]
    norm: str


def normalize(text: str) -> str:
    return re.sub(r"[^a-z0-9]+", " ", text.lower()).strip()


def content_tokens(text: str) -> list[str]:
    tokens = []
    for token in normalize(text).split():
        if token in STOPWORDS:
            continue
        if len(token) < 4 and not any(char.isdigit() for char in token):
            continue
        tokens.append(token)
    return tokens


def unique_tokens(text: str) -> list[str]:
    seen = set()
    tokens = []
    for token in content_tokens(text):
        if token not in seen:
            tokens.append(token)
            seen.add(token)
    return tokens


def field_kind(name: str) -> str:
    lower = name.lower()
    if lower in FINAL_FIELDS:
        return "final"
    if lower.startswith(("answer", "anwser", "clinician recommendation")):
        return "answer"
    return "record"


def workbook_fields(path: Path, include_questions: bool = False) -> tuple[str | None, list[Field]]:
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
    values = [ws.cell(2, col).value for col in range(1, ws.max_column + 1)]
    case_id = str(values[0]).strip() if values and values[0] else None
    fields = []
    for index, header in enumerate(headers):
        if not header:
            continue
        name = str(header).strip()
        value = str(values[index] or "").strip()
        if not value:
            continue
        lower = name.lower()
        if FIELD_RE.fullmatch(name) or lower in FINAL_FIELDS:
            fields.append(Field(name=name, text=value, kind=field_kind(name)))
        elif include_questions and re.fullmatch(r"Question\s+\d+", name, flags=re.IGNORECASE):
            fields.append(Field(name=name, text=value, kind="record"))
    return case_id, fields


def pdf_blocks(doc: fitz.Document) -> list[Block]:
    blocks = []
    for page_index in range(doc.page_count):
        page = doc[page_index]
        for item in page.get_text("blocks"):
            if len(item) >= 7 and item[6] != 0:
                continue
            x0, y0, x1, y1, text = item[:5]
            text = re.sub(r"\s+", " ", str(text)).strip()
            if len(text) < 30:
                continue
            tokens = set(content_tokens(text))
            if not tokens:
                continue
            blocks.append(
                Block(
                    page_index=page_index,
                    rect=fitz.Rect(x0, y0, x1, y1),
                    text=text,
                    tokens=tokens,
                    norm=normalize(text),
                )
            )
    return blocks


def phrase_hits(field_text: str, block_norm: str, size: int = 5) -> int:
    words = normalize(field_text).split()
    if len(words) < size:
        return 0
    hits = 0
    for index in range(0, len(words) - size + 1):
        if " ".join(words[index : index + size]) in block_norm:
            hits += 1
    return hits


def score_block(field: Field, block: Block) -> tuple[float, set[str], int]:
    field_tokens = set(unique_tokens(field.text))
    if not field_tokens:
        return 0.0, set(), 0
    overlap = field_tokens & block.tokens
    if not overlap:
        return 0.0, set(), 0

    field_numbers = {token for token in field_tokens if any(char.isdigit() for char in token)}
    number_overlap = field_numbers & block.tokens
    coverage = len(overlap) / len(field_tokens)
    density = len(overlap) / math.sqrt(max(len(block.tokens), 1))
    phrase_count = phrase_hits(field.text, block.norm)
    number_bonus = len(number_overlap) / max(len(field_numbers), 1) if field_numbers else 0.0
    score = coverage + (0.08 * density) + (0.05 * min(phrase_count, 5)) + (0.15 * number_bonus)
    return score, overlap, phrase_count


def select_blocks(
    field: Field,
    blocks: list[Block],
    max_blocks: int,
    min_score: float,
    min_gain: int,
) -> list[dict]:
    field_tokens = set(unique_tokens(field.text))
    remaining = set(field_tokens)
    candidates = []
    for block in blocks:
        score, overlap, phrase_count = score_block(field, block)
        if score <= 0:
            continue
        candidates.append(
            {
                "block": block,
                "score": score,
                "overlap": overlap,
                "phrase_hits": phrase_count,
            }
        )
    candidates.sort(key=lambda item: item["score"], reverse=True)

    selected = []
    for item in candidates:
        gain = remaining & item["overlap"]
        if item["score"] < min_score and len(gain) < min_gain:
            continue
        selected.append(item)
        remaining -= item["overlap"]
        if len(selected) >= max_blocks:
            break
        if field_tokens and (len(field_tokens) - len(remaining)) / len(field_tokens) >= 0.75:
            break
    return selected


def page_words(page: fitz.Page, page_index: int) -> list[Word]:
    words = []
    for item in page.get_text("words"):
        x0, y0, x1, y1, text, block_no, line_no, word_no = item[:8]
        tokens = content_tokens(str(text))
        words.append(
            Word(
                page_index=page_index,
                rect=fitz.Rect(x0, y0, x1, y1),
                text=str(text),
                tokens=set(tokens),
                block_no=int(block_no),
                line_no=int(line_no),
                word_no=int(word_no),
            )
        )
    return words


def words_in_rect(words: list[Word], rect: fitz.Rect) -> list[Word]:
    return [word for word in words if word.rect.intersects(rect)]


def split_sentences(words: list[Word]) -> list[Sentence]:
    sentences = []
    current = []
    for word in words:
        current.append(word)
        stripped = word.text.strip()
        if re.search(r"[.!?;]$", stripped) and len(current) >= 4:
            text = " ".join(item.text for item in current)
            sentences.append(
                Sentence(
                    page_index=current[0].page_index,
                    words=current,
                    text=text,
                    tokens=set(content_tokens(text)),
                    norm=normalize(text),
                )
            )
            current = []
    if current:
        text = " ".join(item.text for item in current)
        sentences.append(
            Sentence(
                page_index=current[0].page_index,
                words=current,
                text=text,
                tokens=set(content_tokens(text)),
                norm=normalize(text),
            )
        )
    return [sentence for sentence in sentences if sentence.tokens]


def sentence_candidates(
    field: Field,
    selected_blocks: list[dict],
    page_word_cache: dict[int, list[Word]],
) -> list[dict]:
    field_tokens = set(unique_tokens(field.text))
    candidates = []
    if not field_tokens:
        return candidates
    field_numbers = {token for token in field_tokens if any(char.isdigit() for char in token)}
    for item in selected_blocks:
        block = item["block"]
        words = words_in_rect(page_word_cache[block.page_index], block.rect)
        for sentence in split_sentences(words):
            overlap = field_tokens & sentence.tokens
            if not overlap:
                continue
            phrase_count = phrase_hits(field.text, sentence.norm, size=4)
            number_overlap = field_numbers & sentence.tokens
            number_bonus = len(number_overlap) / max(len(field_numbers), 1) if field_numbers else 0.0
            score = (
                len(overlap)
                + (2.0 * min(phrase_count, 3))
                + (2.0 * number_bonus)
                + item["score"]
            )
            candidates.append(
                {
                    "sentence": sentence,
                    "block": block,
                    "score": score,
                    "block_score": item["score"],
                    "overlap": overlap,
                    "phrase_hits": phrase_count,
                }
            )
    candidates.sort(key=lambda item: item["score"], reverse=True)
    return candidates


def select_sentences(
    field: Field,
    selected_blocks: list[dict],
    page_word_cache: dict[int, list[Word]],
    max_sentences: int,
    min_gain: int,
) -> list[dict]:
    field_tokens = set(unique_tokens(field.text))
    remaining = set(field_tokens)
    selected = []
    for item in sentence_candidates(field, selected_blocks, page_word_cache):
        gain = remaining & item["overlap"]
        if len(gain) < min_gain and item["phrase_hits"] == 0:
            continue
        selected.append(item)
        remaining -= item["overlap"]
        if len(selected) >= max_sentences:
            break
        if field_tokens and (len(field_tokens) - len(remaining)) / len(field_tokens) >= 0.75:
            break
    return selected


def keyword_fallback(
    field: Field,
    selected_blocks: list[dict],
    page_word_cache: dict[int, list[Word]],
    limit: int,
) -> list[dict]:
    field_tokens = set(unique_tokens(field.text))
    results = []
    seen = set()
    for item in selected_blocks:
        block = item["block"]
        for word in words_in_rect(page_word_cache[block.page_index], block.rect):
            overlap = word.tokens & field_tokens
            if not overlap:
                continue
            key = (word.page_index, word.block_no, word.line_no, word.word_no)
            if key in seen:
                continue
            seen.add(key)
            results.append(
                {
                    "word": word,
                    "block": block,
                    "score": item["score"],
                    "overlap": overlap,
                }
            )
            if len(results) >= limit:
                return results
    return results


def line_rects(words: list[Word], padding: float = 0.6) -> list[fitz.Rect]:
    lines: dict[tuple[int, int], fitz.Rect] = {}
    for word in words:
        key = (word.block_no, word.line_no)
        rect = fitz.Rect(word.rect)
        if key in lines:
            lines[key].include_rect(rect)
        else:
            lines[key] = rect
    return [
        fitz.Rect(rect.x0 - padding, rect.y0 - padding, rect.x1 + padding, rect.y1 + padding)
        for rect in lines.values()
    ]


def evidence_word_runs(sentence: Sentence, field_tokens: set[str]) -> list[dict]:
    hits = []
    for word in sentence.words:
        overlap = word.tokens & field_tokens
        if overlap:
            hits.append({"word": word, "overlap": overlap})
    if not hits:
        return []

    runs = []
    current_words = [hits[0]["word"]]
    current_overlap = set(hits[0]["overlap"])
    last = hits[0]["word"]
    for hit in hits[1:]:
        word = hit["word"]
        same_line = (word.block_no, word.line_no) == (last.block_no, last.line_no)
        close = word.word_no - last.word_no <= 2
        if same_line and close:
            current_words.append(word)
            current_overlap |= hit["overlap"]
        else:
            runs.append({"words": current_words, "overlap": current_overlap})
            current_words = [word]
            current_overlap = set(hit["overlap"])
        last = word
    runs.append({"words": current_words, "overlap": current_overlap})
    return runs


def add_highlight(
    page: fitz.Page,
    rects: list[fitz.Rect],
    color: tuple[float, float, float],
    label: str,
) -> int:
    if not rects:
        return 0
    annot = page.add_highlight_annot(rects)
    annot.set_colors(stroke=color)
    annot.set_opacity(0.35)
    annot.set_info(content=label)
    annot.update()
    return 1


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", type=Path)
    parser.add_argument("source_pdf", type=Path)
    parser.add_argument("--out", type=Path, required=True)
    parser.add_argument("--report", type=Path)
    parser.add_argument("--max-blocks", type=int, default=3)
    parser.add_argument("--max-sentences-per-field", type=int, default=5)
    parser.add_argument("--min-score", type=float, default=0.12)
    parser.add_argument("--min-gain", type=int, default=4)
    parser.add_argument("--min-sentence-gain", type=int, default=2)
    parser.add_argument("--keyword-fallback-limit", type=int, default=18)
    parser.add_argument("--include-questions", action="store_true")
    parser.add_argument("--allow-empty", action="store_true")
    args = parser.parse_args()

    workbook = args.workbook.expanduser().resolve()
    source_pdf = args.source_pdf.expanduser().resolve()
    out = args.out.expanduser().resolve()
    report_path = args.report.expanduser().resolve() if args.report else None

    case_id, fields = workbook_fields(workbook, include_questions=args.include_questions)
    doc = fitz.open(source_pdf)
    blocks = pdf_blocks(doc)
    page_word_cache = {page_index: page_words(doc[page_index], page_index) for page_index in range(doc.page_count)}

    report = {
        "workbook": str(workbook),
        "source_pdf": str(source_pdf),
        "annotated_pdf": str(out),
        "CRID": case_id,
        "field_count": len(fields),
        "highlight_scope": "sentence_guided_keywords",
        "highlight_count": 0,
        "fields": [],
        "warnings": [],
    }

    for field in fields:
        selected = select_blocks(
            field,
            blocks,
            max_blocks=args.max_blocks,
            min_score=args.min_score,
            min_gain=args.min_gain,
        )
        field_item = {"field": field.name, "kind": field.kind, "matches": []}
        if not selected:
            report["warnings"].append(f"No highlight match selected for {field.name}")
        sentence_matches = select_sentences(
            field,
            selected,
            page_word_cache,
            max_sentences=args.max_sentences_per_field,
            min_gain=args.min_sentence_gain,
        )
        if sentence_matches:
            for item in sentence_matches:
                sentence = item["sentence"]
                page = doc[sentence.page_index]
                run_count = 0
                run_overlaps = set()
                for run in evidence_word_runs(sentence, set(unique_tokens(field.text))):
                    run_count += add_highlight(
                        page,
                        line_rects(run["words"]),
                        COLORS.get(field.kind, COLORS["record"]),
                        f"{case_id or workbook.stem} {field.name} keyword evidence",
                    )
                    run_overlaps |= run["overlap"]
                report["highlight_count"] += run_count
                field_item["matches"].append(
                    {
                        "page": sentence.page_index + 1,
                        "scope": "sentence_guided_keywords",
                        "keyword_annotation_count": run_count,
                        "score": round(item["score"], 3),
                        "block_score": round(item["block_score"], 3),
                        "phrase_hits": item["phrase_hits"],
                        "overlap_tokens": sorted(run_overlaps or item["overlap"])[:30],
                        "text_excerpt": sentence.text[:500],
                    }
                )
        else:
            keyword_matches = keyword_fallback(
                field,
                selected,
                page_word_cache,
                limit=args.keyword_fallback_limit,
            )
            if not keyword_matches and selected:
                report["warnings"].append(f"No sentence or keyword highlight selected for {field.name}")
            for item in keyword_matches:
                word = item["word"]
                page = doc[word.page_index]
                report["highlight_count"] += add_highlight(
                    page,
                    line_rects([word]),
                    COLORS.get(field.kind, COLORS["record"]),
                    f"{case_id or workbook.stem} {field.name} keyword evidence",
                )
                field_item["matches"].append(
                    {
                        "page": word.page_index + 1,
                        "scope": "keyword",
                        "score": round(item["score"], 3),
                        "overlap_tokens": sorted(item["overlap"]),
                        "text_excerpt": word.text,
                    }
                )
        for item in selected:
            block = item["block"]
            field_item.setdefault("source_blocks", []).append(
                {
                    "page": block.page_index + 1,
                    "score": round(item["score"], 3),
                    "phrase_hits": item["phrase_hits"],
                    "overlap_tokens": sorted(item["overlap"])[:30],
                    "text_excerpt": block.text[:300],
                }
            )
        report["fields"].append(field_item)

    out.parent.mkdir(parents=True, exist_ok=True)
    doc.save(out, garbage=4, deflate=True)
    doc.close()

    if report_path:
        report_path.parent.mkdir(parents=True, exist_ok=True)
        report_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")

    print(json.dumps(report, ensure_ascii=False, indent=2))
    if report["highlight_count"] == 0 and not args.allow_empty:
        raise SystemExit(1)


if __name__ == "__main__":
    main()
