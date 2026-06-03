#!/usr/bin/env python3
"""Normalize CR workbooks into compatible case-report JSON."""

from __future__ import annotations

import argparse
import json
import re
from pathlib import Path

from openpyxl import load_workbook


def split_names(value) -> list[dict]:
    if value is None or value == "":
        return []
    out = []
    for name in str(value).split(","):
        name = name.strip()
        if name:
            out.append({"Name": Path(name).name, "Path": name})
    return out


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--out", type=Path, required=True)
    parser.add_argument("--default-question", default="What is the clinician recommendation at this point?")
    args = parser.parse_args()

    wb = load_workbook(args.workbook, data_only=True)
    ws = wb.active
    headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
    values = [ws.cell(2, col).value for col in range(1, ws.max_column + 1)]
    figure_row = [ws.cell(3, col).value for col in range(1, ws.max_column + 1)]
    table_row = [ws.cell(4, col).value for col in range(1, ws.max_column + 1)]

    data: dict = {"CRID": values[0] or args.workbook.stem.split("_")[0]}
    col_by_header = {str(h).strip(): i for i, h in enumerate(headers) if h}

    standard_indices = []
    for header, idx in col_by_header.items():
        match = re.fullmatch(r"Record\s+(\d+)", header)
        if match:
            standard_indices.append((int(match.group(1)), idx))

    if standard_indices:
        for n, record_idx in sorted(standard_indices):
            question_idx = col_by_header.get(f"Question {n}")
            answer_idx = col_by_header.get(f"Answer {n}", col_by_header.get(f"Anwser {n}"))
            block = {
                "Record1": values[record_idx] or "",
                "Figures": split_names(figure_row[record_idx]),
                "Tables": split_names(table_row[record_idx]),
            }
            if question_idx is not None:
                block["Question1"] = values[question_idx] or ""
            if answer_idx is not None:
                block["Answer1"] = values[answer_idx] or ""
            data[f"Consulation{n}"] = block
    else:
        patient_indices = []
        for header, idx in col_by_header.items():
            match = re.fullmatch(r"Patient record\s+(\d+)", header, flags=re.IGNORECASE)
            if match:
                patient_indices.append((int(match.group(1)), idx))
        for n, record_idx in sorted(patient_indices):
            answer_idx = col_by_header.get(f"Clinician recommendation {n}")
            answer = values[answer_idx] if answer_idx is not None else ""
            data[f"Consulation{n}"] = {
                "Record1": values[record_idx] or "",
                "Figures": [],
                "Tables": [],
                "Question1": args.default_question,
                "Answer1": answer or "",
            }

    for header, idx in col_by_header.items():
        if str(header).strip().lower() in {"final follow up", "final output"} and values[idx]:
            data["FinalFollowUp"] = values[idx]

    args.out.parent.mkdir(parents=True, exist_ok=True)
    args.out.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    print(str(args.out))


if __name__ == "__main__":
    main()
