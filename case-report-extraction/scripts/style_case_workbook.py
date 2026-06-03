#!/usr/bin/env python3
"""Apply the canonical case-report workbook style."""

from __future__ import annotations

import argparse
import re
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


FONT_MAIN = Font(name="Calibri", size=16, color="FF000000")
FONT_MAIN_BOLD = Font(name="Calibri", size=16, bold=True, color="FF000000")
FONT_TABLE = Font(name="Calibri", size=11, color="FF000000")

FILL_STAGE_BLUE = PatternFill("solid", fgColor="FFDEEBF7")
FILL_STAGE_GRAY = PatternFill("solid", fgColor="FFF2F2F2")
FILL_FIGURES = PatternFill("solid", fgColor="FFFFF2CC")
FILL_TABLES = PatternFill("solid", fgColor="FFE2F0D9")
NO_FILL = PatternFill(fill_type=None)

THIN_BORDER = Border(
    left=Side(style="thin", color="FF000000"),
    right=Side(style="thin", color="FF000000"),
    top=Side(style="thin", color="FF000000"),
    bottom=Side(style="thin", color="FF000000"),
)
NO_BORDER = Border()

ALIGN_LEFT_WRAP = Alignment(horizontal="left", wrap_text=True)

REFERENCE_WIDTHS = [
    13.0,
    77.14,
    34.86,
    93.71,
    103.29,
    31.57,
    72.86,
    24.71,
    24.0,
    46.29,
    33.71,
    32.0,
    52.57,
    58.71,
    35.0,
    38.0,
    57.86,
    39.71,
    62.86,
    55.0,
    44.57,
    89.86,
    8.71,
    13.0,
    13.0,
    12.43,
    13.0,
    13.0,
    13.0,
    13.0,
    13.0,
    13.0,
    13.0,
    13.0,
]


def set_cell_style(cell, *, font=FONT_MAIN, fill=None, border=NO_BORDER, alignment=ALIGN_LEFT_WRAP):
    cell.font = font
    cell.fill = fill if fill is not None else NO_FILL
    cell.border = border
    cell.alignment = alignment


def style_main_workbook(path: Path) -> None:
    wb = load_workbook(path)
    ws = wb.active
    ws.title = "Sheet1"
    max_col = ws.max_column

    for col in range(1, max_col + 1):
        set_cell_style(ws.cell(1, col), font=FONT_MAIN_BOLD, fill=NO_FILL, border=NO_BORDER)

    for col in range(1, max_col + 1):
        cell = ws.cell(2, col)
        if col == 1:
            set_cell_style(cell, font=FONT_MAIN, fill=NO_FILL, border=NO_BORDER)
        else:
            group_index = (col - 2) // 3
            fill = FILL_STAGE_BLUE if group_index % 2 == 0 else FILL_STAGE_GRAY
            set_cell_style(cell, font=FONT_MAIN, fill=fill, border=THIN_BORDER)

    for col in range(1, max_col + 1):
        set_cell_style(ws.cell(3, col), font=FONT_MAIN, fill=FILL_FIGURES, border=THIN_BORDER)
        set_cell_style(ws.cell(4, col), font=FONT_MAIN, fill=FILL_TABLES, border=THIN_BORDER)

    ws.row_dimensions[1].height = 27.75
    ws.row_dimensions[2].height = 224.25
    ws.row_dimensions[3].height = 27.75
    ws.row_dimensions[4].height = 27.75
    ws.freeze_panes = None

    for col in range(1, max_col + 1):
        letter = get_column_letter(col)
        if col <= len(REFERENCE_WIDTHS):
            width = REFERENCE_WIDTHS[col - 1]
        else:
            offset = (col - 2) % 3
            width = [77.14, 34.86, 93.71][offset]
        ws.column_dimensions[letter].width = width

    wb.save(path)


def style_table_workbook(path: Path) -> None:
    wb = load_workbook(path)
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                set_cell_style(cell, font=FONT_TABLE, fill=NO_FILL, border=NO_BORDER)
        for col in range(1, ws.max_column + 1):
            letter = get_column_letter(col)
            existing = ws.column_dimensions[letter].width
            if existing:
                ws.column_dimensions[letter].width = max(existing, 13.0)
        for row_idx in range(1, ws.max_row + 1):
            ws.row_dimensions[row_idx].height = None
    wb.save(path)


def is_table_workbook(path: Path) -> bool:
    return re.search(r"_table\d+\.xlsx$", path.name, flags=re.IGNORECASE) is not None


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--kind", choices=["auto", "main", "table"], default="auto")
    args = parser.parse_args()

    path = args.workbook.expanduser().resolve()
    kind = args.kind
    if kind == "auto":
        kind = "table" if is_table_workbook(path) else "main"
    if kind == "main":
        style_main_workbook(path)
    else:
        style_table_workbook(path)
    print(path)


if __name__ == "__main__":
    main()
