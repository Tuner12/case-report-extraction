#!/usr/bin/env python3
"""Audit extracted source-table workbooks for structural extraction issues."""

from __future__ import annotations

import argparse
import json
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def table_number(path: Path) -> int:
    match = re.search(r"_table(\d+)\.xlsx$", path.name, flags=re.IGNORECASE)
    return int(match.group(1)) if match else 0


def table_paths(folder: Path, case_id: str) -> list[Path]:
    return sorted(folder.glob(f"{case_id}_table*.xlsx"), key=table_number)


def clean_value(value: Any) -> str:
    if value is None:
        return ""
    return " ".join(str(value).split())


def nonempty_bounds(ws: Any) -> tuple[int, int, int, int] | None:
    min_row = min_col = 10**9
    max_row = max_col = 0
    for row in ws.iter_rows():
        for cell in row:
            if clean_value(cell.value):
                min_row = min(min_row, cell.row)
                max_row = max(max_row, cell.row)
                min_col = min(min_col, cell.column)
                max_col = max(max_col, cell.column)
    if max_row == 0:
        return None
    return min_row, min_col, max_row, max_col


def matrix_values(ws: Any, bounds: tuple[int, int, int, int]) -> list[list[str]]:
    min_row, min_col, max_row, max_col = bounds
    rows: list[list[str]] = []
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        rows.append([clean_value(cell.value) for cell in row])
    return rows


def blank_row_count(rows: list[list[str]]) -> int:
    return sum(1 for row in rows if not any(row))


def blank_col_count(rows: list[list[str]]) -> int:
    if not rows:
        return 0
    return sum(1 for col in range(len(rows[0])) if not any(row[col] for row in rows))


def font_warnings(ws: Any, bounds: tuple[int, int, int, int]) -> list[str]:
    min_row, min_col, max_row, max_col = bounds
    bad_fonts: dict[str, int] = {}
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            if not clean_value(cell.value):
                continue
            font_name = cell.font.name or ""
            if font_name and font_name.lower() != "calibri":
                bad_fonts[font_name] = bad_fonts.get(font_name, 0) + 1
    if not bad_fonts:
        return []
    listed = ", ".join(f"{name} ({count})" for name, count in sorted(bad_fonts.items()))
    return [f"Non-Calibri fonts found in table workbook: {listed}"]


def audit_table(path: Path, args: argparse.Namespace) -> dict[str, Any]:
    warnings: list[str] = []
    try:
        wb = load_workbook(path, data_only=False)
    except Exception as exc:
        return {"table": path.name, "status": "error", "warnings": [f"Cannot open workbook: {exc}"]}

    sheet_reports: list[dict[str, Any]] = []
    for ws in wb.worksheets:
        bounds = nonempty_bounds(ws)
        if bounds is None:
            sheet_reports.append({"sheet": ws.title, "status": "empty", "warnings": ["Sheet has no non-empty cells"]})
            warnings.append(f"{ws.title}: sheet has no non-empty cells")
            continue

        rows = matrix_values(ws, bounds)
        used_rows = len(rows)
        used_cols = len(rows[0]) if rows else 0
        blank_rows = blank_row_count(rows)
        blank_cols = blank_col_count(rows)
        sheet_warnings: list[str] = []
        if used_rows < args.min_rows:
            sheet_warnings.append(f"Only {used_rows} non-empty row(s); verify this is a real source table")
        if used_cols < args.min_cols:
            sheet_warnings.append(f"Only {used_cols} non-empty column(s); verify columns were not collapsed")
        if blank_rows:
            sheet_warnings.append(f"{blank_rows} blank row(s) inside used range")
        if blank_cols:
            sheet_warnings.append(f"{blank_cols} blank column(s) inside used range")
        if list(ws.merged_cells.ranges):
            sheet_warnings.append("Merged cells present; verify headers and row labels were preserved")
        header = rows[0] if rows else []
        if used_rows >= 2 and not any(header):
            sheet_warnings.append("First used row is blank; verify table header alignment")
        sheet_warnings.extend(font_warnings(ws, bounds))

        for warning in sheet_warnings:
            warnings.append(f"{ws.title}: {warning}")
        sheet_reports.append(
            {
                "sheet": ws.title,
                "status": "warning" if sheet_warnings else "ok",
                "used_range": f"{get_column_letter(bounds[1])}{bounds[0]}:{get_column_letter(bounds[3])}{bounds[2]}",
                "used_rows": used_rows,
                "used_cols": used_cols,
                "blank_rows_inside_range": blank_rows,
                "blank_cols_inside_range": blank_cols,
                "merged_ranges": [str(rng) for rng in ws.merged_cells.ranges],
                "warnings": sheet_warnings,
                "preview": rows[: args.preview_rows],
            }
        )

    wb.close()
    return {
        "table": path.name,
        "status": "warning" if warnings else "ok",
        "sheets": sheet_reports,
        "warnings": warnings,
    }


def write_preview(report: dict[str, Any], out: Path) -> None:
    lines: list[str] = []
    for item in report["tables"]:
        lines.append(f"## {item['table']} ({item['status']})")
        for warning in item.get("warnings", []):
            lines.append(f"- WARNING: {warning}")
        for sheet in item.get("sheets", []):
            lines.append(f"### {sheet['sheet']} - {sheet.get('used_range', 'empty')}")
            preview = sheet.get("preview") or []
            if not preview:
                lines.append("(empty)")
                continue
            for row in preview:
                lines.append("| " + " | ".join(cell or " " for cell in row) + " |")
        lines.append("")
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text("\n".join(lines), encoding="utf-8")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("case_folder", type=Path)
    parser.add_argument("--case-id", default=None)
    parser.add_argument("--report", type=Path, default=None)
    parser.add_argument("--preview", type=Path, default=None)
    parser.add_argument("--preview-rows", type=int, default=8)
    parser.add_argument("--min-rows", type=int, default=2)
    parser.add_argument("--min-cols", type=int, default=2)
    parser.add_argument("--require-tables", action="store_true")
    parser.add_argument("--allow-warnings", action="store_true")
    args = parser.parse_args()

    folder = args.case_folder.expanduser().resolve()
    case_id = args.case_id or folder.name
    paths = table_paths(folder, case_id)
    report: dict[str, Any] = {
        "case_folder": str(folder),
        "case_id": case_id,
        "tables_found": len(paths),
        "tables": [],
        "warnings": [],
    }

    if not paths and args.require_tables:
        report["warnings"].append("No table workbooks found, but --require-tables was set")

    for path in paths:
        item = audit_table(path, args)
        report["tables"].append(item)
        for warning in item.get("warnings", []):
            report["warnings"].append(f"{item['table']}: {warning}")

    if args.preview:
        preview_path = args.preview.expanduser().resolve()
        write_preview(report, preview_path)
        report["preview"] = str(preview_path)

    if args.report:
        out = args.report.expanduser().resolve()
        out.parent.mkdir(parents=True, exist_ok=True)
        out.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")

    print(json.dumps(report, ensure_ascii=False, indent=2))
    if report["warnings"] and not args.allow_warnings:
        raise SystemExit(1)


if __name__ == "__main__":
    main()
