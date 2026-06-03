#!/usr/bin/env python3
"""Audit whether workbook fields copy long contiguous spans from the source PDF text."""

from __future__ import annotations

import argparse
import json
import re
from pathlib import Path

from openpyxl import load_workbook


FIELD_RE = re.compile(
    r"^(Record|Answer|Anwser|Patient record|Clinician recommendation)\s+\d+$",
    flags=re.IGNORECASE,
)
FINAL_FIELDS = {"final follow up", "final output"}


def normalize_tokens(text: str) -> list[str]:
    return re.findall(r"[a-z0-9]+", text.lower())


def workbook_fields(path: Path, include_questions: bool = False) -> tuple[str | None, list[tuple[str, str]]]:
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
    values = [ws.cell(2, col).value for col in range(1, ws.max_column + 1)]
    case_id = str(values[0]).strip() if values and values[0] else None
    fields = []
    for index, header in enumerate(headers):
        if not header:
            continue
        name = str(header).strip()
        value = str(values[index] or "").strip()
        if not value:
            continue
        lower = name.lower()
        if FIELD_RE.fullmatch(name) or lower in FINAL_FIELDS:
            fields.append((name, value))
        elif include_questions and re.fullmatch(r"Question\s+\d+", name, flags=re.IGNORECASE):
            fields.append((name, value))
    return case_id, fields


def source_ngram_sets(tokens: list[str], max_n: int, min_n: int) -> dict[int, set[tuple[str, ...]]]:
    sets: dict[int, set[tuple[str, ...]]] = {}
    for size in range(min_n, max_n + 1):
        if len(tokens) < size:
            break
        sets[size] = {tuple(tokens[index : index + size]) for index in range(0, len(tokens) - size + 1)}
    return sets


def longest_source_copy(field_text: str, source_sets: dict[int, set[tuple[str, ...]]], min_words: int) -> dict | None:
    tokens = normalize_tokens(field_text)
    if len(tokens) < min_words:
        return None
    max_size = min(max(source_sets), len(tokens))
    for size in range(max_size, min_words - 1, -1):
        source_ngrams = source_sets.get(size)
        if not source_ngrams:
            continue
        for index in range(0, len(tokens) - size + 1):
            ngram = tuple(tokens[index : index + size])
            if ngram in source_ngrams:
                return {
                    "words": size,
                    "text": " ".join(ngram),
                }
    return None


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--source-text", type=Path, required=True)
    parser.add_argument("--write-report", type=Path)
    parser.add_argument("--max-raw-words", type=int, default=14)
    parser.add_argument("--max-window", type=int, default=80)
    parser.add_argument("--include-questions", action="store_true")
    parser.add_argument("--fail-on-warning", action="store_true")
    args = parser.parse_args()

    source_text = args.source_text.read_text(encoding="utf-8", errors="ignore")
    source_tokens = normalize_tokens(source_text)
    max_window = max(args.max_raw_words, args.max_window)
    source_sets = source_ngram_sets(source_tokens, max_window, args.max_raw_words)
    case_id, fields = workbook_fields(args.workbook, include_questions=args.include_questions)

    report = {
        "workbook": str(args.workbook),
        "source_text": str(args.source_text),
        "CRID": case_id,
        "max_raw_words": args.max_raw_words,
        "fields": [],
        "warnings": [],
        "errors": [],
    }

    for name, text in fields:
        copied = longest_source_copy(text, source_sets, args.max_raw_words)
        item = {"field": name, "status": "ok", "longest_copied_span": copied}
        if copied:
            item["status"] = "warn"
            report["warnings"].append(
                f"{name} contains a {copied['words']}-word contiguous source span"
            )
        report["fields"].append(item)

    if args.write_report:
        args.write_report.parent.mkdir(parents=True, exist_ok=True)
        args.write_report.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")

    print(json.dumps(report, ensure_ascii=False, indent=2))
    if args.fail_on_warning and report["warnings"]:
        raise SystemExit(1)


if __name__ == "__main__":
    main()
