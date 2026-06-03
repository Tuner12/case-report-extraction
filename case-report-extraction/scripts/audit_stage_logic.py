#!/usr/bin/env python3
"""Audit temporal decision logic in a case-report workbook."""

from __future__ import annotations

import argparse
import json
import re
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook


QUESTION_CUES = {
    "what",
    "how",
    "which",
    "whether",
    "should",
    "recommend",
    "recommended",
    "manage",
    "managed",
    "management",
    "treat",
    "treated",
    "treatment",
    "evaluate",
    "evaluation",
    "workup",
    "diagnose",
    "diagnosis",
    "next",
}

ACTION_CUES = {
    "administer",
    "admit",
    "advise",
    "begin",
    "biopsy",
    "continue",
    "discontinue",
    "evaluate",
    "follow",
    "hold",
    "initiate",
    "manage",
    "monitor",
    "obtain",
    "perform",
    "plan",
    "recommend",
    "refer",
    "resume",
    "schedule",
    "send",
    "start",
    "stop",
    "test",
    "treat",
    "withhold",
}

RESULT_CUES = {
    "confirmed",
    "decreased",
    "diagnosis",
    "diagnostic",
    "found",
    "improved",
    "increased",
    "pathology",
    "revealed",
    "showed",
    "shows",
    "result",
    "results",
    "returned",
    "resolved",
    "worsened",
}

FUTURE_CUES = {
    "after discharge",
    "died",
    "death",
    "follow up",
    "follow-up",
    "one week after",
    "recurrence",
    "recurred in",
    "three weeks after",
    "two days later",
}

ACUTE_WORKUP_CUES = {
    "admission",
    "blood pressure",
    "computed tomography",
    "ct",
    "emergency",
    "imaging",
    "laboratory",
    "pulse",
    "respiratory",
    "temperature",
    "ultrasonography",
    "vital",
}

WORKUP_QUESTION_CUES = {
    "diagnostic",
    "evaluation",
    "evaluate",
    "examination",
    "imaging",
    "test",
    "testing",
    "workup",
}

WORKUP_ANSWER_CUES = {
    "blood",
    "culture",
    "ct",
    "examination",
    "imaging",
    "laboratory",
    "microbiologic",
    "panel",
    "pcr",
    "perform",
    "physical",
    "test",
    "testing",
    "ultrasonography",
    "urinalysis",
}

TREATMENT_ANSWER_CUES = {
    "admit",
    "administer",
    "antibiotic",
    "antibiotics",
    "cefepime",
    "fluids",
    "metronidazole",
    "start",
    "treat",
    "vancomycin",
}

LATE_DIAGNOSIS_CUES = {
    "biopsy",
    "confirmed",
    "diagnosis",
    "drug-induced",
    "final",
    "pathological",
    "pathology",
}

DIAGNOSIS_QUESTION_CUES = {
    "cause",
    "diagnose",
    "diagnosis",
    "explain",
    "likely",
}


@dataclass
class Stage:
    number: int
    record: str
    question: str
    answer: str


def normalize(text: str) -> str:
    return re.sub(r"\s+", " ", str(text or "")).strip()


def tokens(text: str) -> set[str]:
    return set(re.findall(r"[a-z0-9-]+", normalize(text).lower()))


def contains_phrase(text: str, phrases: set[str]) -> bool:
    lower = normalize(text).lower()
    return any(phrase in lower for phrase in phrases)


def has_token(text: str, cues: set[str]) -> bool:
    word_set = tokens(text)
    return any(cue in word_set for cue in cues)


def workbook_stages(path: Path) -> tuple[str | None, list[Stage], str]:
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    by_header: dict[str, str] = {}
    case_id = normalize(ws.cell(2, 1).value) or None
    final_follow_up = ""
    for col in range(1, ws.max_column + 1):
        header = normalize(ws.cell(1, col).value)
        value = normalize(ws.cell(2, col).value)
        if not header:
            continue
        by_header[re.sub(r"\s+", " ", header)] = value
        if header.lower() in {"final follow up", "final output"}:
            final_follow_up = value

    stage_numbers = sorted(
        {
            int(match.group(1))
            for header in by_header
            for match in [re.fullmatch(r"Record\s+(\d+)", header, flags=re.IGNORECASE)]
            if match
        }
    )
    stages = []
    for number in stage_numbers:
        stages.append(
            Stage(
                number=number,
                record=by_header.get(f"Record {number}", ""),
                question=by_header.get(f"Question {number}", ""),
                answer=by_header.get(f"Answer {number}", "")
                or by_header.get(f"Anwser {number}", ""),
            )
        )
    return case_id, stages, final_follow_up


def audit_stage(stage: Stage, is_last: bool) -> dict:
    warnings = []
    errors = []

    if not stage.record:
        errors.append("missing Record")
    if not stage.question:
        errors.append("missing Question")
    if not stage.answer and not is_last:
        errors.append("missing Answer")

    if stage.question:
        question_lower = stage.question.lower()
        if "?" not in stage.question and not question_lower.startswith(("what ", "how ", "which ", "should ")):
            warnings.append("Question is not phrased as a decision prompt")
        if not has_token(stage.question, QUESTION_CUES):
            warnings.append("Question lacks clinical decision cues")

    if stage.answer:
        answer_has_action = has_token(stage.answer, ACTION_CUES)
        answer_has_result = has_token(stage.answer, RESULT_CUES)
        question_is_diagnostic = has_token(stage.question, DIAGNOSIS_QUESTION_CUES)
        if not answer_has_action and answer_has_result and not is_last and not question_is_diagnostic:
            warnings.append("Answer reads like a result/diagnosis; consider moving the result into the next Record")
        if not answer_has_action and not is_last and not question_is_diagnostic:
            warnings.append("Answer lacks clear action/recommendation cues")
        if not is_last and contains_phrase(stage.answer, FUTURE_CUES):
            warnings.append("Non-final Answer may contain downstream follow-up or outcome information")
        if (
            has_token(stage.question, WORKUP_QUESTION_CUES)
            and has_token(stage.answer, TREATMENT_ANSWER_CUES)
        ):
            warnings.append(
                "Question asks for evaluation/workup but Answer includes treatment/admission; check whether workup should be this Answer and treatment should move to the next stage"
            )

    if (
        has_token(stage.record, ACUTE_WORKUP_CUES)
        and has_token(stage.answer, LATE_DIAGNOSIS_CUES)
        and not is_last
    ):
        warnings.append(
            "Acute workup Record jumps to late diagnosis in Answer; consider separating immediate management from later diagnostic results"
        )

    return {
        "stage": stage.number,
        "record_chars": len(stage.record),
        "question_chars": len(stage.question),
        "answer_chars": len(stage.answer),
        "warnings": warnings,
        "errors": errors,
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--write-report", type=Path)
    parser.add_argument("--fail-on-warning", action="store_true")
    args = parser.parse_args()

    case_id, stages, final_follow_up = workbook_stages(args.workbook)
    report = {
        "workbook": str(args.workbook),
        "CRID": case_id,
        "stage_count": len(stages),
        "has_final_follow_up": bool(final_follow_up),
        "stages": [],
        "warnings": [],
        "errors": [],
    }

    for index, stage in enumerate(stages):
        item = audit_stage(stage, is_last=index == len(stages) - 1)
        report["stages"].append(item)
        for warning in item["warnings"]:
            report["warnings"].append(f"Stage {stage.number}: {warning}")
        for error in item["errors"]:
            report["errors"].append(f"Stage {stage.number}: {error}")

    if args.write_report:
        args.write_report.parent.mkdir(parents=True, exist_ok=True)
        args.write_report.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")

    print(json.dumps(report, ensure_ascii=False, indent=2))
    if report["errors"] or (args.fail_on_warning and report["warnings"]):
        raise SystemExit(1)


if __name__ == "__main__":
    main()
