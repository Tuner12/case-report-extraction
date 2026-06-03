#!/usr/bin/env python3
"""Convert compatible case-report JSON into the standard wide workbook."""

from __future__ import annotations

import argparse
import json
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


def consultation_items(data: dict) -> list[tuple[int, dict]]:
    items = []
    for key, value in data.items():
        match = re.fullmatch(r"Cons(?:u)?lation(\d+)", key)
        if match and isinstance(value, dict):
            items.append((int(match.group(1)), value))
    if not items and isinstance(data.get("consultations"), list):
        items = [(i + 1, item) for i, item in enumerate(data["consultations"])]
    return sorted(items)


def names(items: list[dict] | None) -> str:
    if not items:
        return ""
    values = []
    for item in items:
        if isinstance(item, dict):
            values.append(str(item.get("Name") or Path(str(item.get("Path", ""))).name))
        else:
            values.append(str(item))
    return ", ".join([v for v in values if v])


def field(block: dict, *keys: str) -> str:
    for key in keys:
        if block.get(key):
            return str(block[key])
    return ""


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("json_file", type=Path)
    parser.add_argument("--out", type=Path, required=True)
    parser.add_argument(
        "--modern-answer-headers",
        action="store_true",
        help="Use Answer 4 instead of legacy Anwser 4.",
    )
    args = parser.parse_args()

    data = json.loads(args.json_file.read_text(encoding="utf-8"))
    case_id = data.get("CRID") or data.get("case_id") or args.json_file.stem.split("_")[0]
    items = consultation_items(data)
    if not items:
        raise SystemExit("No ConsulationN or consultations[] blocks found")

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    headers = ["CRID"]
    values = [case_id]
    figures = ["figures"]
    tables = ["tables"]

    final_follow_up = data.get("FinalFollowUp") or data.get("Final follow up") or data.get("Final output")

    for index, block in items:
        record = field(block, "Record1", "record", "Patient record")
        question = field(block, "Question1", "question")
        answer = field(block, "Answer1", "answer", "Clinician recommendation")
        block_final = field(block, "FinalFollowUp", "Final follow up", "Final output")

        headers.append(f"Record {index}")
        values.append(record)
        figures.append(names(block.get("Figures") or block.get("figures")))
        tables.append(names(block.get("Tables") or block.get("tables")))

        if question or answer:
            headers.append(f"Question {index}")
            values.append(question)
            figures.append("")
            tables.append("")
            if index == 4 and not args.modern_answer_headers:
                headers.append("Anwser 4")
            else:
                headers.append(f"Answer {index}")
            values.append(answer)
            figures.append("")
            tables.append("")
        elif block_final:
            headers.append("Final follow up")
            values.append(block_final)
            figures.append("")
            tables.append("")

    if final_follow_up:
        headers.append("Final follow up")
        values.append(str(final_follow_up))
        figures.append("")
        tables.append("")

    for row_idx, row in enumerate([headers, values, figures, tables], start=1):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        width = 14 if col == 1 else 36
        ws.column_dimensions[letter].width = width
    ws.row_dimensions[2].height = 180
    ws.freeze_panes = "B2"

    args.out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.out)
    print(str(args.out))


if __name__ == "__main__":
    main()
